"""Export a completed trading session to an Excel workbook.

Reads JSONL trade/order files produced by TradeStore and writes one .xlsx with:
  - Trades            — one row per closed trade
  - Orders            — full order audit log
  - OrderStagesSummary — counts by stage × status
  - ExitReasonSummary  — counts + net P&L by exit reason
"""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any


def _load_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with open(path, encoding="utf-8") as fh:
        for raw in fh:
            line = raw.strip()
            if line:
                try:
                    rows.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return rows


def export_session_excel(
    engine_name: str,
    trade_day: str,
    base_dir: str = "trade_store",
    out_dir: str = "Results/SessionReports",
) -> str | None:
    """Build the Excel report and return the output path, or None on failure."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        from logger import log_event
        log_event("[Report] openpyxl not installed — skipping Excel export", "warning")
        return None

    base = Path(base_dir)
    trades = _load_jsonl(base / f"{engine_name}_{trade_day}_trades.jsonl")
    orders = _load_jsonl(base / f"{engine_name}_{trade_day}_orders.jsonl")

    if not trades and not orders:
        from logger import log_event
        log_event("[Report] No trade/order data found — skipping Excel export", "info")
        return None

    wb = openpyxl.Workbook()

    # ── helpers ───────────────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="1E293B")
    HEADER_FONT = Font(bold=True, color="94A3B8")

    def _make_sheet(name: str, columns: list[str], rows: list[list]) -> None:
        ws = wb.create_sheet(name)
        ws.append(columns)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # ── Trades sheet ──────────────────────────────────────────────────────────
    trade_cols = [
        "trade_id", "symbol", "side", "quantity",
        "entry_time", "exit_time", "entry_price", "exit_price",
        "pnl", "estimated_charges", "net_pnl", "pnl_pct",
        "exit_reason", "engine_name", "execution_mode", "pair_id", "recorded_at",
    ]
    trade_rows = [[r.get(c) for c in trade_cols] for r in trades]
    _make_sheet("Trades", trade_cols, trade_rows)

    # ── Orders sheet ──────────────────────────────────────────────────────────
    order_cols = [
        "audit_id", "stage", "symbol", "side", "quantity",
        "product", "execution_mode", "provider", "status",
        "order_id", "entry_price", "message", "note", "recorded_at",
    ]
    order_rows = [[r.get(c) for c in order_cols] for r in orders]
    _make_sheet("Orders", order_cols, order_rows)

    # ── OrderStagesSummary sheet ──────────────────────────────────────────────
    stage_counts: dict[tuple[str, str], int] = defaultdict(int)
    for r in orders:
        stage_counts[(r.get("stage", ""), r.get("status", ""))] += 1
    summary_rows = [[stage, status, cnt] for (stage, status), cnt in sorted(stage_counts.items())]
    _make_sheet("OrderStagesSummary", ["stage", "status", "count"], summary_rows)

    # ── ExitReasonSummary sheet ───────────────────────────────────────────────
    reason_data: dict[str, dict] = defaultdict(lambda: {"count": 0, "net_pnl": 0.0, "wins": 0})
    for r in trades:
        reason = r.get("exit_reason") or "UNKNOWN"
        net = float(r.get("net_pnl") or 0)
        reason_data[reason]["count"] += 1
        reason_data[reason]["net_pnl"] += net
        if net > 0:
            reason_data[reason]["wins"] += 1
    exit_rows = [
        [reason, d["count"], d["wins"], round(d["net_pnl"], 2)]
        for reason, d in sorted(reason_data.items(), key=lambda x: -x[1]["count"])
    ]
    _make_sheet("ExitReasonSummary", ["exit_reason", "count", "wins", "net_pnl"], exit_rows)

    # ── remove default blank sheet ────────────────────────────────────────────
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── save ──────────────────────────────────────────────────────────────────
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{engine_name}_{trade_day}_session_report_{ts}.xlsx"
    dest = out_path / file_name
    wb.save(dest)
    return str(dest)
