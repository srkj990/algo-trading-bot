from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
import csv
import re
import zipfile
from xml.sax.saxutils import escape


def summarize_by_exit_reason(trade_book: list[dict]) -> list[dict]:
    buckets: dict[str, dict] = {}
    for trade in trade_book or []:
        reason = str(trade.get("exit_reason") or "UNKNOWN")
        bucket = buckets.setdefault(
            reason,
            {
                "exit_reason": reason,
                "trades": 0,
                "gross_pnl": 0.0,
                "net_pnl": 0.0,
                "wins": 0,
                "losses": 0,
                "flats": 0,
            },
        )
        pnl = float(trade.get("pnl") or 0.0)
        net = float(trade.get("net_pnl") or pnl)
        bucket["trades"] += 1
        bucket["gross_pnl"] += pnl
        bucket["net_pnl"] += net
        if pnl > 0:
            bucket["wins"] += 1
        elif pnl < 0:
            bucket["losses"] += 1
        else:
            bucket["flats"] += 1

    rows = list(buckets.values())
    rows.sort(key=lambda r: (r["gross_pnl"], r["trades"]), reverse=True)
    for row in rows:
        trades = max(1, int(row["trades"]))
        row["avg_gross_pnl"] = row["gross_pnl"] / trades
        row["avg_net_pnl"] = row["net_pnl"] / trades
        row["win_rate"] = (row["wins"] / trades) * 100.0
    return rows


def _ensure_results_dir() -> Path:
    results_dir = Path("Results")
    results_dir.mkdir(exist_ok=True)
    return results_dir


_COL_RE = re.compile(r"^[A-Z]+$")


def _excel_col_name(n: int) -> str:
    # 1-indexed -> A, B, ..., Z, AA, AB, ...
    name = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        name = chr(ord("A") + rem) + name
    return name


def _xlsx_cell_xml(value, cell_ref: str) -> str:
    if value is None:
        return f'<c r="{cell_ref}"/>'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f'<c r="{cell_ref}" t="n"><v>{value}</v></c>'
    text = escape(str(value))
    return f'<c r="{cell_ref}" t="inlineStr"><is><t>{text}</t></is></c>'


def _write_simple_xlsx(path: Path, *, sheets: list[tuple[str, list[list]]]) -> None:
    """
    Write a minimal, dependency-free .xlsx with inline strings.
    Enough for Excel/LibreOffice to open and for humans to read.
    """

    # Sheet XMLs
    sheet_files = []
    for idx, (name, rows) in enumerate(sheets, start=1):
        sheet_name = escape(name)
        xml_rows = []
        for r_idx, row in enumerate(rows, start=1):
            cells = []
            for c_idx, value in enumerate(row, start=1):
                cell_ref = f"{_excel_col_name(c_idx)}{r_idx}"
                cells.append(_xlsx_cell_xml(value, cell_ref))
            xml_rows.append(f'<row r="{r_idx}">' + "".join(cells) + "</row>")
        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            "<sheetData>"
            + "".join(xml_rows)
            + "</sheetData></worksheet>"
        )
        sheet_path = f"xl/worksheets/sheet{idx}.xml"
        sheet_files.append((sheet_path, sheet_xml))

    workbook_sheets = []
    workbook_rels = []
    for idx, (name, _rows) in enumerate(sheets, start=1):
        sheet_name = escape(name)
        workbook_sheets.append(
            f'<sheet name="{sheet_name}" sheetId="{idx}" r:id="rId{idx}"/>'
        )
        workbook_rels.append(
            (
                f'<Relationship Id="rId{idx}" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
                f'Target="worksheets/sheet{idx}.xml"/>'
            )
        )

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        "<sheets>"
        + "".join(workbook_sheets)
        + "</sheets></workbook>"
    )

    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + "".join(workbook_rels)
        + "</Relationships>"
    )

    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        + "".join(
            f'<Override PartName="/xl/worksheets/sheet{idx}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            for idx in range(1, len(sheets) + 1)
        )
        + "</Types>"
    )

    root_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        "</Relationships>"
    )

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml)
        zf.writestr("_rels/.rels", root_rels_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        for sheet_path, sheet_xml in sheet_files:
            zf.writestr(sheet_path, sheet_xml)


def export_trade_book_report(trade_book: list[dict], *, engine_name: str) -> Path | None:
    if not trade_book:
        return None

    results_dir = _ensure_results_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = results_dir / f"{engine_name}_trade_report_{timestamp}"

    # Try Excel first (openpyxl). Fallback to CSV if dependency is unavailable.
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        path = base.with_suffix(".xlsx")
        wb = openpyxl.Workbook()

        ws_trades = wb.active
        ws_trades.title = "Trades"

        # Stable column ordering (human-friendly)
        preferred_cols = [
            "symbol",
            "side",
            "quantity",
            "entry_time",
            "exit_time",
            "entry_price",
            "exit_price",
            "pnl",
            "estimated_charges",
            "net_pnl",
            "pnl_pct",
            "exit_reason",
            "pair_id",
        ]
        all_cols = set()
        for t in trade_book:
            all_cols.update(t.keys())
        extra_cols = [c for c in sorted(all_cols) if c not in preferred_cols]
        columns = preferred_cols + extra_cols

        ws_trades.append(columns)
        for trade in trade_book:
            ws_trades.append([trade.get(col) for col in columns])

        for col_idx, col_name in enumerate(columns, start=1):
            width = max(10, min(40, len(str(col_name)) + 2))
            ws_trades.column_dimensions[get_column_letter(col_idx)].width = width

        ws_summary = wb.create_sheet("ExitReasonSummary")
        summary_rows = summarize_by_exit_reason(trade_book)
        summary_cols = [
            "exit_reason",
            "trades",
            "gross_pnl",
            "net_pnl",
            "avg_gross_pnl",
            "avg_net_pnl",
            "wins",
            "losses",
            "flats",
            "win_rate",
        ]
        ws_summary.append(summary_cols)
        for row in summary_rows:
            ws_summary.append([row.get(col) for col in summary_cols])
        for col_idx, col_name in enumerate(summary_cols, start=1):
            width = max(12, min(32, len(str(col_name)) + 2))
            ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(path)
        return path
    except Exception:
        # Dependency-free XLSX (preferred), then CSV if something goes wrong.
        try:
            path = base.with_suffix(".xlsx")

            preferred_cols = [
                "symbol",
                "side",
                "quantity",
                "entry_time",
                "exit_time",
                "entry_price",
                "exit_price",
                "pnl",
                "estimated_charges",
                "net_pnl",
                "pnl_pct",
                "exit_reason",
                "pair_id",
            ]
            all_cols = set()
            for t in trade_book:
                all_cols.update(t.keys())
            extra_cols = [c for c in sorted(all_cols) if c not in preferred_cols]
            columns = preferred_cols + extra_cols

            trades_rows = [columns]
            for trade in trade_book:
                trades_rows.append([trade.get(col) for col in columns])

            summary_rows = summarize_by_exit_reason(trade_book)
            summary_cols = [
                "exit_reason",
                "trades",
                "gross_pnl",
                "net_pnl",
                "avg_gross_pnl",
                "avg_net_pnl",
                "wins",
                "losses",
                "flats",
                "win_rate",
            ]
            summary_table = [summary_cols]
            for row in summary_rows:
                summary_table.append([row.get(col) for col in summary_cols])

            _write_simple_xlsx(
                path,
                sheets=[
                    ("Trades", trades_rows),
                    ("ExitReasonSummary", summary_table),
                ],
            )
            return path
        except Exception:
            # CSV fallback (Excel can still open it)
            path = base.with_suffix(".csv")
            summary_path = base.with_name(base.name + "_exit_summary").with_suffix(".csv")

            cols = sorted({k for t in trade_book for k in t.keys()})
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=cols)
                writer.writeheader()
                for trade in trade_book:
                    writer.writerow(trade)

            summary_rows = summarize_by_exit_reason(trade_book)
            with open(summary_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=summary_rows[0].keys() if summary_rows else ["exit_reason"],
                )
                writer.writeheader()
                for row in summary_rows:
                    writer.writerow(row)

            return path
